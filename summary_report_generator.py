"""
美团经营宝汇总报表生成器

将所有板块的报表合并到单个 Excel 文件中，每个板块一个 Sheet。

使用方法：
    python summary_report_generator.py

输出：
    config/reports/美团经营宝汇总_YYYY-MM-DD.xlsx
"""

import os
import sys
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 获取项目根目录
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = _CURRENT_DIR
sys.path.insert(0, _PROJECT_ROOT)


# ============== 样式定义 ==============

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
DATA_FONT = Font(name="微软雅黑", size=10)

TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")  # 橙色
HEADER_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFFFE6E6", end_color="FFFFE6E6", fill_type="solid")
FLOW_FILL = PatternFill(start_color="FFE3F1D9", end_color="FFE3F1D9", fill_type="solid")  # 浅绿（简版看板基础信息）
COMPETE_FILL = PatternFill(start_color="FFE3F2D9", end_color="FFE3F2D9", fill_type="solid")  # 浅绿（推广通竞争分析）
FLOW_TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")  # 橙色（简版看板流量数据）
STAR_FILL = PatternFill(start_color="FFD9E8FF", end_color="FFD9E8FF", fill_type="solid")  # 浅蓝（简版看板星级数据）
REVIEW_FILL = PatternFill(start_color="FFD9D0FF", end_color="FFD9D0FF", fill_type="solid")  # 浅紫（简版看板门店评价）
# 客流分析分组颜色
FLOW_USER_FILL = PatternFill(start_color="FFD6E8FF", end_color="FFD6E8FF", fill_type="solid")  # 浅蓝（引流用户数据）
ONLINE_CONSULT_FILL = PatternFill(start_color="FFD5F5DC", end_color="FFD5F5DC", fill_type="solid")  # 浅绿（在线咨询）
TRADE_FILL = PatternFill(start_color="FFF0E68C", end_color="FFF0E68C", fill_type="solid")  # 浅黄（门店交易）
REVIEW_OVERVIEW_FILL = PatternFill(start_color="FFEEE8F5", end_color="FFEEE8F5", fill_type="solid")  # 浅紫（门店评价概览）
STAR_OVERVIEW_FILL = PatternFill(start_color="FFFCE4EC", end_color="FFFCE4EC", fill_type="solid")  # 浅粉（星级概览）

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


class SummaryExcelGenerator:
    """汇总报表生成器"""

    def __init__(self):
        self.wb = Workbook()
        # 删除默认 sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    def add_sheet_from_generator(self, sheet_name: str, generator, title: str, headers: list, widths: list, data_rows: list, row_heights: list = None, title_style: str = "default"):
        """
        从已有 ExcelGenerator 的数据创建 Sheet

        Args:
            sheet_name: Sheet 名称
            generator: ExcelGenerator 实例（用于获取样式）
            title: 标题
            headers: 表头列表
            widths: 列宽列表
            data_rows: 数据行列表（每行是列表）
            row_heights: 每行的行高列表，默认 [28, 22, 18, ...]
            title_style: 标题样式
                - "default": 统一样式（橙色背景）
                - "tuiguangtong": 推广通样式（A1:R1橙色 + S1:V1无背景"竞争分析"）
                - "simple_board": 简版看板样式（A1:L1浅绿 + M1:P1橙色 + Q1:R1浅蓝 + S1:Y1浅紫）
                - "customer_flow": 客流分析样式（多分组）
        """
        ws = self.wb.create_sheet(title=sheet_name)

        # 确定列数
        num_cols = max(len(headers), len(widths), 1)
        if data_rows:
            num_cols = max(num_cols, len(data_rows[0]))

        # 行高设置：默认 [标题行=28, 表头行=22, 数据行=18]
        if row_heights is None:
            row_heights = [28, 22] + [18] * len(data_rows) if data_rows else [28, 22]

        # 标题行
        last_col = get_column_letter(num_cols) if num_cols > 0 else "A"

        if title_style == "tuiguangtong" and num_cols >= 22:
            # 推广通样式：A1:R1 橙色标题 + S1:V1 "竞争分析"（浅绿色）
            ws.merge_cells("A1:R1")
            title_cell = ws["A1"]
            title_cell.value = title
            title_cell.font = TITLE_FONT
            title_cell.fill = TITLE_FILL
            title_cell.alignment = CENTER_ALIGN

            ws.merge_cells("S1:V1")
            comp_cell = ws["S1"]
            comp_cell.value = "竞争分析"
            comp_cell.font = TITLE_FONT
            comp_cell.fill = COMPETE_FILL
            comp_cell.alignment = CENTER_ALIGN

        elif title_style == "simple_board" and num_cols >= 25:
            # 简版看板样式：4个分组
            # A1:L1 浅绿 + M1:P1 橙色 + Q1:R1 浅蓝 + S1:Y1 浅紫
            ws.merge_cells("A1:L1")
            title_cell = ws["A1"]
            title_cell.value = title
            title_cell.font = TITLE_FONT
            title_cell.fill = FLOW_FILL  # 浅绿
            title_cell.alignment = CENTER_ALIGN

            ws.merge_cells("M1:P1")
            flow_cell = ws["M1"]
            flow_cell.value = "流量数据"
            flow_cell.font = TITLE_FONT
            flow_cell.fill = FLOW_TITLE_FILL  # 橙色
            flow_cell.alignment = CENTER_ALIGN

            ws.merge_cells("Q1:R1")
            star_cell = ws["Q1"]
            star_cell.value = "星级数据"
            star_cell.font = TITLE_FONT
            star_cell.fill = STAR_FILL  # 浅蓝
            star_cell.alignment = CENTER_ALIGN

            ws.merge_cells("S1:Y1")
            review_cell = ws["S1"]
            review_cell.value = "门店评价"
            review_cell.font = TITLE_FONT
            review_cell.fill = REVIEW_FILL  # 浅紫
            review_cell.alignment = CENTER_ALIGN

        elif title_style == "customer_flow" and num_cols >= 39:
            # 客流分析样式：6个分组
            # A1:P1 客流统计(橙色) + Q1:S1 引流用户数据(浅蓝) + T1:V1 在线咨询(浅绿)
            # W1:AF1 门店交易情况(浅黄) + AG1:AK1 门店评价概览(浅紫) + AL1:AM1 星级概览(浅粉)
            ws.merge_cells("A1:P1")
            title_cell = ws["A1"]
            title_cell.value = title
            title_cell.font = TITLE_FONT
            title_cell.fill = TITLE_FILL  # 橙色
            title_cell.alignment = CENTER_ALIGN

            ws.merge_cells("Q1:S1")
            user_cell = ws["Q1"]
            user_cell.value = "引流用户数据情况"
            user_cell.font = TITLE_FONT
            user_cell.fill = FLOW_USER_FILL  # 浅蓝
            user_cell.alignment = CENTER_ALIGN

            ws.merge_cells("T1:V1")
            consult_cell = ws["T1"]
            consult_cell.value = "在线咨询"
            consult_cell.font = TITLE_FONT
            consult_cell.fill = ONLINE_CONSULT_FILL  # 浅绿
            consult_cell.alignment = CENTER_ALIGN

            ws.merge_cells("W1:AF1")
            trade_cell = ws["W1"]
            trade_cell.value = "门店交易情况"
            trade_cell.font = TITLE_FONT
            trade_cell.fill = TRADE_FILL  # 浅黄
            trade_cell.alignment = CENTER_ALIGN

            ws.merge_cells("AG1:AK1")
            review_cell = ws["AG1"]
            review_cell.value = "门店评价概览"
            review_cell.font = TITLE_FONT
            review_cell.fill = REVIEW_OVERVIEW_FILL  # 浅紫
            review_cell.alignment = CENTER_ALIGN

            ws.merge_cells("AL1:AM1")
            star_cell = ws["AL1"]
            star_cell.value = "星级概览"
            star_cell.font = TITLE_FONT
            star_cell.fill = STAR_OVERVIEW_FILL  # 浅粉
            star_cell.alignment = CENTER_ALIGN

        elif title_style == "simple_board_old" and num_cols >= 13:
            # 旧版简版看板样式（简版看板报表 - A1:I1 绿色，J-M 橙色）
            ws.merge_cells(f"A1:I1")
            title_cell = ws["A1"]
            title_cell.value = title
            title_cell.font = TITLE_FONT
            title_cell.fill = FLOW_FILL  # 绿色
            title_cell.alignment = CENTER_ALIGN

            ws.merge_cells(f"J1:{last_col}1")
            flow_cell = ws["J1"]
            flow_cell.value = "流量数据"
            flow_cell.font = TITLE_FONT
            flow_cell.fill = TITLE_FILL  # 橙色
            flow_cell.alignment = CENTER_ALIGN
        else:
            # 默认统一标题样式
            ws.merge_cells(f"A1:{last_col}1")
            title_cell = ws["A1"]
            title_cell.value = title
            title_cell.font = TITLE_FONT
            title_cell.fill = TITLE_FILL
            title_cell.alignment = CENTER_ALIGN

        ws.row_dimensions[1].height = row_heights[0] if len(row_heights) > 0 else 28

        # 表头行
        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = row_heights[1] if len(row_heights) > 1 else 22

        # 数据行
        for row_idx, row_data in enumerate(data_rows, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.fill = WHITE_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
            # 使用传入的行高，第3行对应 row_heights[2]，以此类推
            height_idx = row_idx - 1  # row_idx=3 -> idx=2
            ws.row_dimensions[row_idx].height = row_heights[height_idx] if height_idx < len(row_heights) else 18

    def add_error_sheet(self, sheet_name: str, module_name: str, error_msg: str):
        """添加错误提示 Sheet"""
        ws = self.wb.create_sheet(title=sheet_name)

        # 标题
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"{module_name} - 数据获取失败"
        title_cell.font = TITLE_FONT
        title_cell.fill = ERROR_FILL
        title_cell.alignment = CENTER_ALIGN
        ws.row_dimensions[1].height = 28

        # 错误信息
        ws["A2"] = "错误信息"
        ws["A2"].font = HEADER_FONT
        ws["A2"].fill = HEADER_FILL
        ws["A2"].border = THIN_BORDER
        ws["A2"].alignment = CENTER_ALIGN

        ws["B2"] = error_msg
        ws["B2"].font = DATA_FONT
        ws["B2"].fill = WHITE_FILL
        ws["B2"].border = THIN_BORDER
        ws["B2"].alignment = LEFT_ALIGN

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 60

    def save(self, filepath: str):
        """保存 Excel 文件"""
        # 检查文件是否被占用
        if os.path.exists(filepath):
            try:
                with open(filepath, 'a'):
                    pass
            except IOError:
                print(f"错误: 文件被占用，请关闭 '{filepath}' 后再重试")
                raise IOError(f"文件被占用: {filepath}")
        self.wb.save(filepath)
        print(f"汇总报表已保存: {filepath}")


def load_main_config():
    """加载主配置文件"""
    config_path = os.path.join(_PROJECT_ROOT, "config", "main_config.json")
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def get_shop_name(shop_id: str) -> str:
    """从门店映射表获取门店名称"""
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    if shop_id == "0":
        return "全部门店汇总数据"

    for shop in shops:
        for id_info in shop.get("ids", []):
            if id_info.get("id") == shop_id:
                return shop.get("name", f"门店{shop_id}")
    return f"门店{shop_id}"


def get_shop_city(shop_id: str) -> str:
    """从门店映射表获取门店所在城市"""
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    if shop_id == "0":
        return ""

    for shop in shops:
        for id_info in shop.get("ids", []):
            if id_info.get("id") == shop_id:
                return shop.get("city", "")
    return ""


# ============== 各板块数据获取 ==============

def generate_tuiguangtong_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """推广通汇总数据"""
    from tuiguangtong.client import TuiguangtongClient
    from tuiguangtong.excel_generator import ExcelGenerator

    client = TuiguangtongClient()
    generator = ExcelGenerator()
    generator._create_title_row("推广通报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        city = get_shop_city(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            # tuiguangtong get_metrics: (begin_date, end_date, shop_ids="0", platform=0, ...)
            metrics = client.get_metrics(
                begin_date=begin_date,
                end_date=end_date,
                shop_ids=shop_id,  # 单个 shop_id 字符串
                platform=platform
            )
            # tuiguangtong add_data_row: (shop_id, shop_name, city, date_range, investment_days, metrics, competition)
            generator.add_data_row(shop_id, shop_name, city, date_range, 0, metrics, {})
            print("成功")
            success_count += 1
        except Exception as e:
            error_msg = str(e)
            if "noRightsShop" in error_msg:
                print(f"无权限访问，已跳过")
            else:
                print(f"失败 - {e}")
                generator.add_data_row(shop_id, shop_name, city, date_range, 0, [], {})
                fail_count += 1

    return generator, success_count, fail_count


def generate_review_statistics_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """评论统计汇总数据"""
    from review_statistics.client import ReviewStatisticsClient
    from review_statistics.excel_generator import ExcelGenerator

    client = ReviewStatisticsClient()
    generator = ExcelGenerator()
    generator._create_title_row("评论统计报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        platform_name = {0: "全平台", 1: "点评", 2: "美团"}.get(platform, str(platform))
        print(f"  [{shop_name}]...", end=" ")
        try:
            result = client.get_statistics(shop_id, begin_date, end_date, str(platform))
            if result.get("code") == 200:
                statistics = client.parse_statistics(result)
                generator.add_data_row(shop_id, shop_name, platform_name, date_range, statistics)
                print(f"成功 - 累计评价 {statistics.get('累计评价数', 0)}")
                success_count += 1
            else:
                print(f"失败 - {result.get('msg')}")
                generator.add_data_row(shop_id, shop_name, platform_name, date_range, {})
                fail_count += 1
        except Exception as e:
            print(f"失败 - {e}")
            generator.add_data_row(shop_id, shop_name, platform_name, date_range, {})
            fail_count += 1

    return generator, success_count, fail_count


def generate_star_rating_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """星级评分汇总数据"""
    from star_rating.client import StarRatingClient
    from star_rating.excel_generator import ExcelGenerator

    client = StarRatingClient()
    generator = ExcelGenerator()
    generator._create_title_row("星级评分报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            result = client.get_statistics(shop_id, begin_date, end_date, str(platform))
            if result.get("code") == 200:
                statistics = client.parse_statistics(result)
                generator.add_data_row(shop_id, shop_name, date_range, statistics)
                print(f"成功 - 点评{statistics.get('点评星级', '-')}星, 美团{statistics.get('美团星级', '-')}星")
                success_count += 1
            else:
                print(f"失败 - {result.get('msg')}")
                generator.add_data_row(shop_id, shop_name, date_range, {})
                fail_count += 1
        except Exception as e:
            print(f"失败 - {e}")
            generator.add_data_row(shop_id, shop_name, date_range, {})
            fail_count += 1

    return generator, success_count, fail_count


def generate_shop_review_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """门店评论汇总数据"""
    from shop_review.client import ShopReviewClient
    from shop_review.excel_generator import ExcelGenerator

    client = ShopReviewClient()
    generator = ExcelGenerator()
    generator._create_title_row("门店评论报表")
    generator._create_header_row()

    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            reviews = client.get_reviews(shop_id, str(platform), begin_date, end_date)
            # 获取平台名称
            platform_name = {0: "全平台", 1: "点评", 2: "美团"}.get(platform, str(platform))
            for review in reviews:
                detail = client.get_review_detail(review)
                generator.add_review_row(detail, platform_name)
            print(f"成功({len(reviews)}条)")
            success_count += 1
        except Exception as e:
            print(f"失败 - {e}")
            fail_count += 1

    return generator, success_count, fail_count


def generate_online_consultation_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """在线咨询汇总数据"""
    from online_consultation.client import OnlineConsultationClient
    from online_consultation.excel_generator import ExcelGenerator

    client = OnlineConsultationClient()
    generator = ExcelGenerator()
    generator._create_title_row("在线咨询分析报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            metrics = client.get_metrics(
                shop_id=shop_id,
                platform=str(platform),
                begin_date=begin_date,
                end_date=end_date
            )
            generator.add_data_row(shop_id, shop_name, date_range, metrics)
            print("成功")
            success_count += 1
        except Exception as e:
            print(f"失败 - {e}")
            generator.add_data_row(shop_id, shop_name, date_range, {})
            fail_count += 1

    return generator, success_count, fail_count


def generate_trade_analysis_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """交易分析汇总数据"""
    from trade_analysis.client import TradeAnalysisClient
    from trade_analysis.excel_generator import ExcelGenerator

    client = TradeAnalysisClient()
    generator = ExcelGenerator()
    generator._create_title_row("交易分析报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            metrics = client.get_metrics(
                shop_id=shop_id,
                begin_date=begin_date,
                end_date=end_date,
                platform=str(platform)
            )
            generator.add_data_row(shop_id, shop_name, date_range, metrics)
            print("成功")
            success_count += 1
        except Exception as e:
            print(f"失败 - {e}")
            generator.add_data_row(shop_id, shop_name, date_range, {})
            fail_count += 1

    return generator, success_count, fail_count


def generate_customer_flow_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """客流统计汇总数据"""
    from customer_flow.client import CustomerFlowClient
    from customer_flow.excel_generator import ExcelGenerator

    # flow_metrics 和 user_data_metrics 的 key 列表
    FLOW_METRIC_KEYS = [
        "exposure_count", "exposure_times", "visitor_count", "visit_times",
        "exposure_visit_rate", "intention_count", "intention_rate",
        "order_count", "retention_count", "favorites_count",
        "new_favorites_count", "new_checkin_count"
    ]
    USER_DATA_METRIC_KEYS = ["traffic_user_count", "natural_user_count", "potential_user_count"]

    client = CustomerFlowClient()
    generator = ExcelGenerator()
    generator._create_title_row()
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        city = get_shop_city(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            metrics = client.get_metrics(
                shop_id=shop_id,
                begin_date=begin_date,
                end_date=end_date,
                platform=platform
            )
            # 从扁平字典中提取 flow_metrics 和 user_data_metrics
            flow_metrics = {k: metrics.get(k, "") for k in FLOW_METRIC_KEYS}
            user_data_metrics = {k: metrics.get(k, "") for k in USER_DATA_METRIC_KEYS}
            generator.add_data_row(
                shop_id, shop_name, city, date_range,
                flow_metrics, user_data_metrics, {}, {}, {}, {}
            )
            print("成功")
            success_count += 1
        except Exception as e:
            error_msg = str(e)
            if "noRightsShop" in error_msg:
                print(f"无权限访问，已跳过")
            else:
                print(f"失败 - {e}")
                generator.add_data_row(shop_id, shop_name, city, date_range, {}, {}, {}, {}, {}, {})
                fail_count += 1

    return generator, success_count, fail_count


def generate_quantianzhan_summary(shop_ids: list, begin_date: str, end_date: str, platform: int):
    """全站推广汇总数据"""
    from quantianzhan.client import QuantianzhanClient
    from quantianzhan.excel_generator import ExcelGenerator

    client = QuantianzhanClient()
    generator = ExcelGenerator()
    generator._create_title_row("全站推广报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        city = get_shop_city(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            metrics = client.get_metrics(
                begin_date=begin_date,
                end_date=end_date,
                platform=str(platform),
                shop_ids=shop_id
            )
            generator.add_data_row(shop_id, shop_name, city, date_range, 0, metrics)
            print("成功")
            success_count += 1
        except Exception as e:
            error_msg = str(e)
            if "noRightsShop" in error_msg:
                print(f"无权限访问，已跳过")
            else:
                print(f"失败 - {e}")
                generator.add_data_row(shop_id, shop_name, city, date_range, 0, [])
                fail_count += 1

    return generator, success_count, fail_count


def generate_simple_board_summary(shop_ids: list, begin_date: str, end_date: str, platform: int, review_platform: int = 2, date_scope: int = 2):
    """简版看板汇总数据"""
    from simple_board.client import SimpleBoardClient
    from simple_board.excel_generator import ExcelGenerator

    client = SimpleBoardClient()
    generator = ExcelGenerator()
    generator._create_title_row()
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    success_count = 0
    fail_count = 0

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        city = get_shop_city(shop_id)
        print(f"  [{shop_name}]...", end=" ")
        try:
            metrics = client.get_metrics(
                begin_date=begin_date,
                end_date=end_date,
                shop_id=shop_id,
                platform=str(platform)
            )
            # 获取评论统计数据
            review_stats = client.get_review_statistics(
                shop_id=shop_id,
                review_platform=str(review_platform),
                date_scope=date_scope
            )
            platform_name = "美团" if review_platform == 2 else "点评"
            date_scope_text = {1: "昨天", 2: "近7天", 3: "近30天"}.get(date_scope, "近7天")
            generator.add_data_row(shop_id, shop_name, city, date_range, metrics, platform_name, date_scope_text, review_stats)
            print("成功")
            success_count += 1
        except Exception as e:
            error_msg = str(e)
            if "noRightsShop" in error_msg:
                print(f"无权限访问，已跳过")
            else:
                print(f"失败 - {e}")
                generator.add_data_row(shop_id, shop_name, city, date_range, {}, "美团", "近7天", {})
                fail_count += 1

    return generator, success_count, fail_count


# ============== 主流程 ==============

def generate_summary_report(config: dict = None) -> str:
    """
    生成汇总报表（4个Sheet：推广通、全站推广、简版看板、客流分析）

    Args:
        config: 可选，配置字典。如果为 None，则从 main_config.json 加载。
               格式: {
                   "门店列表": ["id1", "id2"],
                   "平台": 0,
                   "评论统计平台": 1,
                   "评论统计日期范围": 2,
                   "日期范围": {"begin": "2026-01-01", "end": "2026-03-31"}
               }

    Returns:
        str: 生成的文件路径
    """
    print("=" * 60)
    print("美团经营宝汇总报表生成器")
    print("=" * 60)

    # 加载配置
    if config is None:
        config = load_main_config()
    shop_ids = config.get("门店列表", [])
    platform = config.get("平台", 0)
    review_platform = config.get("评论统计平台", 2)
    date_scope = config.get("评论统计日期范围", 2)
    date_range = config.get("日期范围", {})
    begin_date = date_range.get("begin")
    end_date = date_range.get("end")

    print(f"\n配置信息:")
    print(f"  门店数量: {len(shop_ids)}")
    print(f"  平台: {platform}")
    print(f"  评论统计平台: {review_platform}")
    print(f"  评论统计日期范围: {date_scope}")
    print(f"  日期: {begin_date} 至 {end_date}")
    print()

    # 创建汇总生成器
    summary_generator = SummaryExcelGenerator()

    # 定义板块顺序和生成函数（只保留4个）
    modules = [
        ("推广通", generate_tuiguangtong_summary),
        ("全站推广", generate_quantianzhan_summary),
        ("简版看板", generate_simple_board_summary),
        ("付费版线上数据", generate_customer_flow_summary),
    ]

    # 遍历各板块
    for module_name, generate_func in modules:
        print(f"\n[{module_name}]")
        try:
            # 根据板块特性调用不同的参数
            if module_name == "简版看板":
                generator, success_count, fail_count = generate_func(
                    shop_ids, begin_date, end_date, platform, review_platform, date_scope
                )
            else:
                generator, success_count, fail_count = generate_func(
                    shop_ids, begin_date, end_date, platform
                )
            # 从生成器获取数据（统一从worksheet读取）
            headers = generator._get_headers() if hasattr(generator, '_get_headers') else []
            widths = generator._get_widths() if hasattr(generator, '_get_widths') else []
            data_rows = []
            for row in generator.ws.iter_rows(min_row=3, values_only=True):
                if any(cell is not None for cell in row):
                    data_rows.append(list(row))

            # 根据板块特性设置 title_style
            title_style_map = {
                "推广通": "tuiguangtong",
                "简版看板": "simple_board",
                "付费版线上数据": "customer_flow",
            }
            title_style = title_style_map.get(module_name, "default")

            summary_generator.add_sheet_from_generator(
                sheet_name=module_name,
                generator=generator,
                title=f"{module_name}报表",
                headers=headers,
                widths=widths,
                data_rows=data_rows,
                title_style=title_style
            )
            print(f"  → 添加 Sheet 成功 (成功:{success_count}, 失败:{fail_count})")
        except Exception as e:
            print(f"  → 添加 Sheet 失败: {e}")
            summary_generator.add_error_sheet(module_name, module_name, str(e))

    # 保存文件
    output_dir = os.path.join(_PROJECT_ROOT, "config", "reports")
    os.makedirs(output_dir, exist_ok=True)

    # 判断是多门店还是单门店
    is_multi_shop = len(shop_ids) > 1
    shop_label = "多门店" if is_multi_shop else shop_ids[0]
    date_str = f"{begin_date}_{end_date}"
    output_file = os.path.join(output_dir, f"美团经营宝汇总_{shop_label}_{date_str}.xlsx")

    summary_generator.save(output_file)

    print("\n" + "=" * 60)
    print("汇总报表生成完成")
    print("=" * 60)

    return output_file


if __name__ == "__main__":
    generate_summary_report()
