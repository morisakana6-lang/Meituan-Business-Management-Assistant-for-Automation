"""
简版看板多门店ID测试脚本
使用实际payload参数测试多个门店ID，生成汇总报表
"""

import sys
import os

_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)

from simple_board.client import SimpleBoardClient


def test_multi_shop_report():
    """测试多个门店ID，生成汇总报表"""
    print("=" * 60)
    print("简版看板多门店ID测试")
    print("=" * 60)

    # 使用两个门店ID测试
    shop_ids = ["1732110850", "1027064882511291"]
    # 使用配置中的日期
    begin_date = "2026-05-01"
    end_date = "2026-05-10"
    platform = "0"

    print(f"\n测试门店IDs: {shop_ids}")
    print(f"日期范围: {begin_date} 至 {end_date}")
    print(f"平台: {platform}")

    client = SimpleBoardClient()

    # 从shop_mapping获取门店名称
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    def get_shop_name(shop_id: str) -> str:
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}")
        return f"门店{shop_id}"

    results = []

    # 测试1: 分别查询每个门店
    print("\n" + "=" * 60)
    print("分别查询每个门店")
    print("=" * 60)

    for shop_id in shop_ids:
        print(f"\n正在查询门店 {shop_id}...")

        try:
            metrics = client.get_metrics(
                begin_date=begin_date,
                end_date=end_date,
                shop_id=shop_id,
                platform=platform
            )

            shop_name = get_shop_name(shop_id)
            results.append({
                "shop_id": shop_id,
                "shop_name": shop_name,
                "metrics": metrics,
            })
            print(f"  {shop_name}: 获取成功")
            print(f"    概览: {metrics.get('overview')}")
            print(f"    交易: {metrics.get('trade')}")
            print(f"    流量: {metrics.get('flow')}")

        except Exception as e:
            print(f"  请求异常: {e}")

    # 测试2: 使用逗号分隔的多门店ID
    print("\n" + "=" * 60)
    print("使用逗号分隔的多门店ID查询")
    print("=" * 60)

    multi_shop_id = ",".join(shop_ids)
    print(f"\nshop_id: {multi_shop_id}")

    try:
        metrics_multi = client.get_metrics(
            begin_date=begin_date,
            end_date=end_date,
            shop_id=multi_shop_id,
            platform=platform
        )
        print(f"  获取成功")
        print(f"    概览: {metrics_multi.get('overview')}")
        print(f"    交易: {metrics_multi.get('trade')}")
        print(f"    流量: {metrics_multi.get('flow')}")
    except Exception as e:
        print(f"  请求异常: {e}")

    # 测试3: 使用shop_id=0全部门店汇总
    print("\n" + "=" * 60)
    print("使用shop_id=0全部门店汇总")
    print("=" * 60)

    try:
        metrics_all = client.get_metrics(
            begin_date=begin_date,
            end_date=end_date,
            shop_id="0",
            platform=platform
        )
        print(f"  获取成功")
        print(f"    概览: {metrics_all.get('overview')}")
        print(f"    交易: {metrics_all.get('trade')}")
        print(f"    流量: {metrics_all.get('flow')}")
    except Exception as e:
        print(f"  请求异常: {e}")

    # 生成Excel报表
    if results:
        generate_excel_report(results, begin_date, end_date)


def generate_excel_report(results, begin_date, end_date):
    """生成Excel报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
    DATA_FONT = Font(name="微软雅黑", size=10)

    TITLE_FILL = PatternFill(start_color="FFE3F1D9", end_color="FFE3F1D9", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "简版看板报表"

    # 表头
    headers = [
        "门店ID", "推广门店", "时间",
        "曝光人数", "访问人数", "下单券数",
        "下单金额(原价)", "核销券数", "核销金额(原价)",
        "曝光次数", "曝光人数", "访问次数", "访问人数"
    ]

    widths = [21.125, 28.875, 25.375, 10.0, 13.0, 13.0, 14.0, 10.0, 14.0, 10.0, 13.0, 13.0, 13.0]

    # 标题行
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value = "简版看板报表（多门店）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28

    # 表头行
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # 数据行
    date_range = f"{begin_date}至{end_date}"

    def format_value(value):
        """格式化数值，移除逗号"""
        if isinstance(value, str):
            return value.replace(",", "")
        return value

    for row_idx, result in enumerate(results, start=3):
        shop_id = result["shop_id"]
        shop_name = result["shop_name"]
        metrics = result["metrics"]

        overview = metrics.get("overview", {})
        trade = metrics.get("trade", {})
        flow = metrics.get("flow", {})

        row_data = [
            shop_id,
            shop_name,
            date_range,
            format_value(overview.get("exposure_count", "0")),
            format_value(overview.get("visitor_count", "0")),
            format_value(trade.get("order_count", "0")),
            format_value(trade.get("order_amount", "0")),
            format_value(trade.get("verify_count", "0")),
            format_value(trade.get("verify_amount", "0")),
            format_value(flow.get("exposure_times", "0")),
            format_value(flow.get("exposure_count", "0")),
            format_value(flow.get("visit_times", "0")),
            format_value(flow.get("visitor_count", "0")),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        ws.row_dimensions[row_idx].height = 18

    # 保存
    output_dir = os.path.join(_CURRENT_DIR, "reports")
    os.makedirs(output_dir, exist_ok=True)
    output_file = f"{output_dir}/简版看板_多门店_{begin_date}_{end_date}.xlsx"

    # 检查文件是否被占用
    if os.path.exists(output_file):
        try:
            with open(output_file, 'a'):
                pass
        except IOError:
            print(f"错误: 文件被占用 {output_file}")
            print("请关闭 Excel 文件后重试")
            raise IOError(f"文件被占用: {output_file}")

    wb.save(output_file)
    print(f"\nExcel 文件已保存: {output_file}")


if __name__ == "__main__":
    test_multi_shop_report()
