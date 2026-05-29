"""
简版看板Excel生成器

=== 板块归属 ===
简版看板板块，对应目录 simple_board/

=== 样式说明 ===
- 第1行: 大标题，简版看板(D-I列)橙色背景，流量数据(J-M列)橙色背景
- 第2行: 表头，灰色背景
- 第3行起: 数据行，白色背景
- 字体: 微软雅黑
- 数值右对齐，文本左对齐
"""

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict

import sys
import os

# 获取项目根目录
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)


# ============== 样式定义 ==============

# 字体
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)  # 大标题
SECTION_FONT = Font(name="微软雅黑", bold=True, size=11)  # 分组标题
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)  # 表头
DATA_FONT = Font(name="微软雅黑", size=10)  # 数据

# 填充色
TITLE_FILL = PatternFill(start_color="FFE3F1D9", end_color="FFE3F1D9", fill_type="solid")  # 浅绿色（基础数据标题行）
FLOW_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")  # 浅橙色（流量数据标题行）
STAR_FILL = PatternFill(start_color="D9E8FF", end_color="D9E8FF", fill_type="solid")  # 浅蓝色（星级数据标题行）
REVIEW_FILL = PatternFill(start_color="D9D0FF", end_color="D9D0FF", fill_type="solid")  # 浅紫色（门店评价标题行）
HEADER_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")  # 灰色（表头行）
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # 白色（数据行）

# 对齐
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


class ExcelGenerator:
    """简版看板Excel生成器"""

    def __init__(self, account_name: str = "简版看板"):
        self.account_name = account_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "简版看板报表"

    def _create_title_row(self, title: str = None):
        """创建标题行（第1行）"""
        # 基础数据分组标题 (A1:L1) - 浅绿色
        self.ws.merge_cells("A1:L1")
        title_cell = self.ws["A1"]
        title_cell.value = title if title else "简版看板"
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL  # 浅绿色
        title_cell.alignment = CENTER_ALIGN

        # 流量数据分组标题 (M1:P1) - 浅橙色
        self.ws.merge_cells("M1:P1")
        flow_cell = self.ws["M1"]
        flow_cell.value = "流量数据"
        flow_cell.font = TITLE_FONT
        flow_cell.fill = FLOW_FILL  # 浅橙色
        flow_cell.alignment = CENTER_ALIGN

        # 星级数据分组标题 (Q1:R1) - 浅蓝色
        self.ws.merge_cells("Q1:R1")
        star_cell = self.ws["Q1"]
        star_cell.value = "星级数据"
        star_cell.font = TITLE_FONT
        star_cell.fill = STAR_FILL  # 浅蓝色
        star_cell.alignment = CENTER_ALIGN

        # 门店评价分组标题 (S1:Y1) - 浅紫色
        self.ws.merge_cells("S1:Y1")
        review_cell = self.ws["S1"]
        review_cell.value = "门店评价"
        review_cell.font = TITLE_FONT
        review_cell.fill = REVIEW_FILL  # 浅紫色
        review_cell.alignment = CENTER_ALIGN

        # 设置第1行行高
        self.ws.row_dimensions[1].height = 21.0

    def _create_header_row(self):
        """创建表头行（第2行）"""
        headers = self._get_headers()
        widths = self._get_widths()

        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = self.ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 设置第2行行高
        self.ws.row_dimensions[2].height = 16.5

    def _get_widths(self):
        """获取列宽列表"""
        return [
            25.375,  # A 门店ID
            27.875,  # B 推广门店
            13.0,    # C 门店所在城市
            23.375,  # D 时间
            7.875,   # E 曝光人数
            13.0,    # F 访问人数
            25.625,  # G 访问率
            7.875,   # H 下单券数
            12.625,  # I 下单金额(原价)
            7.875,   # J 核销券数
            25.625,  # K 核销率
            12.625,  # L 核销金额(原价)
            7.875,   # M 曝光次数
            13.0,    # N 曝光人数
            13.0,    # O 访问次数
            13.0,    # P 访问人数
            11.25,   # Q 大众点评星级
            7.875,   # R 美团星级
            8.0,     # S 平台
            25.0,    # T 统计周期
            9.0,     # U 累计评价数
            13.0,    # V 新增评价数
            13.0,    # W 新增好评数
            13.0,    # X 新增差评数
            13.0,    # Y 差评回复率
        ]

    def _get_headers(self):
        """获取表头列表"""
        return [
            "门店ID",           # A
            "推广门店",         # B
            "门店所在城市",     # C
            "时间",             # D
            "曝光人数",         # E
            "访问人数",         # F
            "访问率",           # G
            "下单券数",         # H
            "下单金额(原价)",   # I
            "核销券数",         # J
            "核销率",           # K
            "核销金额(原价)",   # L
            "曝光次数",         # M
            "曝光人数",         # N
            "访问次数",         # O
            "访问人数",         # P
            "大众点评星级",     # Q
            "美团星级",         # R
            "平台",             # S
            "统计周期",         # T
            "累计评价数",       # U
            "新增评价数",       # V
            "新增好评数",       # W
            "新增差评数",       # X
            "差评回复率",       # Y
        ]

    def _format_value(self, value: str) -> str:
        """格式化数值，移除逗号"""
        if isinstance(value, str):
            return value.replace(",", "")
        return value

    def _parse_float(self, value: str) -> float:
        """解析数值为浮点数，支持万、亿单位"""
        if isinstance(value, str):
            value = value.replace(",", "").replace("%", "").replace(" ", "")
            # 处理万、亿单位
            if "亿" in value:
                value = value.replace("亿", "")
                try:
                    return float(value) * 100000000
                except (ValueError, TypeError):
                    return 0.0
            elif "万" in value:
                value = value.replace("万", "")
                try:
                    return float(value) * 10000
                except (ValueError, TypeError):
                    return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _calculate_rate(self, numerator: str, denominator: str) -> str:
        """计算比率，返回百分比字符串"""
        num = self._parse_float(numerator)
        den = self._parse_float(denominator)
        if den > 0:
            return f"{num / den * 100:.1f}%"
        return "0.0%"

    def add_data_row(
        self,
        shop_id: str,
        shop_name: str,
        city: str,
        date_str: str,
        metrics: Dict,
        platform_name: str = "美团",
        review_date_range: str = "",
        review_stats: Dict = None,
    ):
        """
        添加一行数据

        Args:
            shop_id: 门店ID
            shop_name: 门店名称
            city: 门店所在城市
            date_str: 时间描述（日期范围）
            metrics: 指标字典，包含 overview, trade, flow, star 四类数据
            platform_name: 平台名称（美团/点评）
            review_date_range: 评论统计的统计周期
            review_stats: 评论统计数据字典
        """
        if review_stats is None:
            review_stats = {}

        overview = metrics.get("overview", {})
        trade = metrics.get("trade", {})
        flow = metrics.get("flow", {})
        star = metrics.get("star", {})

        # 计算访问率和核销率
        exposure_count = overview.get("exposure_count", "0")
        visitor_count = overview.get("visitor_count", "0")
        order_count = trade.get("order_count", "0")
        verify_count = trade.get("verify_count", "0")

        visit_rate = self._calculate_rate(visitor_count, exposure_count)
        verify_rate = self._calculate_rate(verify_count, order_count)

        row_data = [
            shop_id,                                              # A 门店ID
            shop_name,                                            # B 推广门店
            city,                                                 # C 门店所在城市
            date_str,                                             # D 时间
            self._format_value(exposure_count),                   # E 曝光人数
            self._format_value(visitor_count),                    # F 访问人数
            visit_rate,                                           # G 访问率
            self._format_value(order_count),                      # H 下单券数
            self._format_value(trade.get("order_amount", "0")),   # I 下单金额(原价)
            self._format_value(verify_count),                     # J 核销券数
            verify_rate,                                          # K 核销率
            self._format_value(trade.get("verify_amount", "0")),  # L 核销金额(原价)
            self._format_value(flow.get("exposure_times", "0")),  # M 曝光次数
            self._format_value(flow.get("exposure_count", "0")),  # N 曝光人数
            self._format_value(flow.get("visit_times", "0")),     # O 访问次数
            self._format_value(flow.get("visitor_count", "0")),   # P 访问人数
            star.get("dp_star", "-"),                             # Q 大众点评星级
            star.get("mt_star", "-"),                             # R 美团星级
            platform_name,                                        # S 平台
            review_date_range,                                    # T 统计周期
            review_stats.get("累计评价数", "-"),                 # U 累计评价数
            review_stats.get("新增评价数", "-"),                 # V 新增评价数
            review_stats.get("新增好评数", "-"),                 # W 新增好评数
            review_stats.get("新增差评数", "-"),                 # X 新增差评数
            review_stats.get("差评回复率", "-"),                 # Y 差评回复率
        ]

        row_idx = self.ws.max_row + 1

        for col_idx, value in enumerate(row_data, start=1):
            cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER

                    # 全部居中对齐
            cell.alignment = CENTER_ALIGN

        # 设置数据行行高
        self.ws.row_dimensions[row_idx].height = 16.5

    def save(self, filepath: str):
        """保存Excel文件"""
        # 检查文件是否被占用
        if os.path.exists(filepath):
            try:
                with open(filepath, 'a'):
                    pass
            except IOError:
                print(f"错误: 文件被占用，请关闭 '{filepath}' 后再重试")
                raise IOError(f"文件被占用: {filepath}")
        self.wb.save(filepath)
        print(f"Excel 文件已保存: {filepath}")


def generate_report(
    client,
    shop_ids: list,
    begin_date: str,
    end_date: str,
    platform: int = 2,
    review_platform: int = 2,
    date_scope: int = 2,
) -> str:
    """
    生成简版看板报表（支持单门店和多门店）

    Args:
        client: SimpleBoardClient 实例
        shop_ids: 门店ID列表，支持单个或多个门店
        begin_date: 开始日期，格式 YYYY-MM-DD
        end_date: 结束日期，格式 YYYY-MM-DD
        platform: 平台选择，1=点评，2=美团
        review_platform: 评论统计平台，1=点评，2=美团
        date_scope: 评论统计日期范围，1=昨天，2=近7天，3=近30天

    Returns:
        生成的 Excel 文件路径
    """
    from datetime import datetime, timedelta

    # 加载门店映射表
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    def get_shop_info(shop_id: str) -> tuple:
        """获取门店名称和城市"""
        if shop_id == "0":
            return "全部门店汇总数据", ""
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}"), shop.get("city", "")
        return f"门店{shop_id}", ""

    # 根据 dateScope 计算评论统计的日期范围
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    if date_scope == 1:
        review_date_begin = yesterday.strftime("%Y-%m-%d")
        review_date_end = review_date_begin
    elif date_scope == 2:
        review_date_begin = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        review_date_end = yesterday.strftime("%Y-%m-%d")
    else:  # date_scope == 3
        review_date_begin = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        review_date_end = yesterday.strftime("%Y-%m-%d")
    review_date_range = f"{review_date_begin}至{review_date_end}"

    platform_name = "美团" if platform == 2 else "点评"

    generator = ExcelGenerator()

    # 判断单门店还是多门店
    is_multi_shop = len(shop_ids) > 1
    if is_multi_shop:
        generator._create_title_row("简版看板报表（多门店）")
    else:
        generator._create_title_row()

    generator._create_header_row()

    # 日期范围
    date_str = f"{begin_date}至{end_date}"
    print(f"开始生成简版看板报表: {begin_date} ~ {end_date}")
    print(f"评论统计: 平台={platform_name}, 统计周期={review_date_range}")

    # 遍历每个门店查询数据
    for shop_id in shop_ids:
        shop_name, city = get_shop_info(shop_id)
        print(f"  查询门店: {shop_name}...")

        metrics = {}
        review_stats = {}

        # 获取简版看板指标数据
        metrics_failed = False
        try:
            metrics = client.get_metrics(
                begin_date=begin_date,
                end_date=end_date,
                shop_id=shop_id,
                platform=str(platform)
            )
            print(f"    简版看板指标: 获取成功")
        except Exception as e:
            error_msg = str(e)
            if "noRightsShop" in error_msg:
                print(f"    简版看板指标: 无权限访问，已跳过")
                metrics_failed = True
            else:
                print(f"    简版看板指标: 获取失败 - {e}")
                metrics_failed = True

        # 如果简版看板指标获取失败，跳过该门店
        if metrics_failed:
            continue

        # 获取评论统计数据
        try:
            review_stats = client.get_review_statistics(
                shop_id=shop_id,
                review_platform=str(review_platform),
                date_scope=date_scope
            )
            if review_stats:
                print(f"    评论统计数据: 获取成功")
            else:
                print(f"    评论统计数据: 无数据")
        except Exception as e:
            print(f"    评论统计数据: 获取失败 - {e}")

        print(f"    {shop_name}: 获取成功")

        generator.add_data_row(shop_id, shop_name, city, date_str, metrics, platform_name, review_date_range, review_stats)

    # 保存文件
    output_dir = os.path.join(_CURRENT_DIR, "reports")
    os.makedirs(output_dir, exist_ok=True)

    if is_multi_shop:
        output_file = f"{output_dir}/简版看板_多门店_{begin_date}_{end_date}.xlsx"
    else:
        output_file = f"{output_dir}/简版看板_{shop_ids[0]}_{begin_date}_{end_date}.xlsx"

    generator.save(output_file)
    return output_file


if __name__ == "__main__":
    """
    入口：直接运行 python excel_generator.py 即可生成报表

    使用方法：
        1. 修改 shop_config.json 中的 search_key（门店ID或名称）和日期
        2. 运行 python excel_generator.py

    search_key: 可以是单个门店ID、多个门店ID列表（JSON数组格式）或逗号分隔字符串
    platform: 0=全平台（默认），1=点评，2=美团
    """
    import json
    from client import SimpleBoardClient

    # 复用推广通的门店搜索模块
    _SHOP_SEARCH_PATH = os.path.join(_PROJECT_ROOT, 'tuiguangtong', 'shop_search.py')
    sys.path.insert(0, os.path.dirname(_SHOP_SEARCH_PATH))
    from shop_search import is_shop_id, search_by_id, resolve_shop

    # 读取配置
    config_path = os.path.join(os.path.dirname(__file__), "shop_config.json")
    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    search_key = config.get("search_key", "")
    platform = config.get("platform", 2)
    review_platform = config.get("reviewPlatform", 2)
    date_scope = config.get("dateScope", 2)
    date_range = config["日期范围"]

    # 解析 search_key（支持多种格式）
    if isinstance(search_key, list):
        shop_ids = search_key
    elif isinstance(search_key, str):
        if "," in search_key:
            shop_ids = [s.strip() for s in search_key.split(",")]
        else:
            shop_ids = [search_key]
    else:
        shop_ids = [str(search_key)]

    # 获取门店信息
    if len(shop_ids) == 1:
        if shop_ids[0] == "0":
            shop_name = "全部门店汇总数据"
        elif is_shop_id(str(shop_ids[0])):
            matches = search_by_id(str(shop_ids[0]))
            if matches:
                shop_name = matches[0]["name"]
            else:
                shop_name = f"门店ID:{shop_ids[0]}"
        else:
            shop_name, _ = resolve_shop(str(shop_ids[0]), platform)
        print(f"已选择门店: {shop_name}")
    else:
        first_key = shop_ids[0]
        if is_shop_id(str(first_key)):
            matches = search_by_id(str(first_key))
            if matches:
                shop_name = matches[0]["name"]
            else:
                shop_name = f"门店ID:{first_key}"
        else:
            shop_name, _ = resolve_shop(str(first_key), platform)
        print(f"已选择门店数量: {len(shop_ids)}")

    platform_names = {1: "点评", 2: "美团"}
    print(f"平台: {platform_names.get(platform, '美团')}")
    print(f"评论统计平台: {platform_names.get(review_platform, '美团')}")
    print(f"统计周期: dateScope={date_scope}")
    print()

    client = SimpleBoardClient()

    generate_report(
        client=client,
        shop_ids=shop_ids,
        begin_date=date_range["begin"],
        end_date=date_range["end"],
        platform=platform,
        review_platform=review_platform,
        date_scope=date_scope
    )