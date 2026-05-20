"""
评论统计Excel生成器

=== 板块归属 ===
评论统计板块，对应目录 review_statistics/

=== 样式说明 ===
- 第1行: 大标题，橙色背景
- 第2行: 表头，灰色背景
- 第3行起: 数据行，白色背景
- 字体: 微软雅黑
- 全部居中对齐
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, List

import sys
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)


# ============== 样式定义 ==============

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
DATA_FONT = Font(name="微软雅黑", size=10)

TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


class ExcelGenerator:
    """评论统计Excel生成器"""

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "评论统计"

    def _create_title_row(self, title: str = "评论统计报表"):
        """创建标题行（第1行）"""
        headers = self._get_headers()
        last_col = get_column_letter(len(headers))
        self.ws.merge_cells(f"A1:{last_col}1")
        title_cell = self.ws["A1"]
        title_cell.value = title
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = CENTER_ALIGN
        self.ws.row_dimensions[1].height = 28

    def _create_header_row(self):
        """创建表头行（第2行）"""
        headers = self._get_headers()
        widths = self._get_widths()

        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = self.ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width

        self.ws.row_dimensions[2].height = 22

    def _get_headers(self):
        """获取表头列表"""
        return [
            "门店ID",
            "门店名称",
            "统计周期",
            "累计评价数",
            "新增评价数",
            "新增好评数",
            "新增差评数",
            "差评回复率",
        ]

    def _get_widths(self):
        """获取列宽列表"""
        return [15, 30, 25, 15, 15, 15, 15, 15]

    def _get_metric_value_float(self, value: str) -> float:
        """将值转换为浮点数"""
        if isinstance(value, str):
            # 先移除逗号、货币符号、百分号和空格
            value = value.replace(",", "").replace("￥", "").replace("¥", "").replace("$", "").replace("%", "").replace(" ", "")
            # 处理带"万"、"亿"的单位
            if "万" in value:
                value = value.replace("万", "")
                try:
                    return float(value) * 10000
                except (ValueError, TypeError):
                    return 0.0
            elif "亿" in value:
                value = value.replace("亿", "")
                try:
                    return float(value) * 100000000
                except (ValueError, TypeError):
                    return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def add_data_row(self, shop_id: str, shop_name: str, date_range: str, statistics: Dict):
        """
        添加一行统计数据

        Args:
            shop_id: 门店ID
            shop_name: 门店名称
            date_range: 统计周期
            statistics: 统计数据字典
        """
        row_data = [
            shop_id,
            shop_name,
            date_range,
            self._get_metric_value_float(statistics.get("累计评价数", "0")),
            self._get_metric_value_float(statistics.get("新增评价数", "0")),
            self._get_metric_value_float(statistics.get("新增好评数", "0")),
            self._get_metric_value_float(statistics.get("新增差评数", "0")),
            statistics.get("差评回复率", ""),
        ]

        row_idx = self.ws.max_row + 1

        for col_idx, value in enumerate(row_data, start=1):
            cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        self.ws.row_dimensions[row_idx].height = 18

    def save(self, filepath: str):
        """保存Excel文件"""
        # 检查文件是否被占用
        if os.path.exists(filepath):
            try:
                with open(filepath, 'a'):
                    pass
            except IOError:
                print(f"错误: 文件被占用 {filepath}")
                print("请关闭 Excel 文件后重试")
                raise IOError(f"文件被占用: {filepath}")
        self.wb.save(filepath)
        print(f"Excel 文件已保存: {filepath}")


def generate_report(
    client,
    shop_ids: List[str],
    platform: int = 0,
    begin_date: str = None,
    end_date: str = None,
    output_dir: str = None
) -> str:
    """
    生成评论统计报表（支持多门店）

    Args:
        client: ReviewStatisticsClient 实例
        shop_ids: 门店ID列表
        platform: 平台选择，0=全平台，1=点评，2=美团
        begin_date: 开始日期，格式 YYYY-MM-DD
        end_date: 结束日期，格式 YYYY-MM-DD
        output_dir: 输出目录，默认 D:\code\meituan\review_statistics\reports

    Returns:
        生成的 Excel 文件路径
    """
    # 加载门店映射表
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    def get_shop_name(shop_id: str) -> str:
        if shop_id == "0":
            return "全部门店汇总数据"
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}")
        return f"门店{shop_id}"

    # 设置默认日期
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    if not begin_date:
        begin_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    # 判断是否多门店
    is_multi_shop = len(shop_ids) > 1

    # 创建Excel
    generator = ExcelGenerator()
    if is_multi_shop:
        generator._create_title_row("评论统计报表（多门店）")
    else:
        generator._create_title_row("评论统计报表")
    generator._create_header_row()

    date_range = f"{begin_date}至{end_date}"
    print(f"开始生成评论统计报表: {begin_date} ~ {end_date}")

    # 遍历每个门店获取数据
    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        print(f"  查询门店: {shop_name}...")

        try:
            result = client.get_statistics(shop_id, begin_date, end_date, str(platform))
            if result.get("code") == 200:
                statistics = client.parse_statistics(result)
                generator.add_data_row(shop_id, shop_name, date_range, statistics)
                print(f"    {shop_name}: 获取成功 - 累计评价 {statistics.get('累计评价数', 0)}")
            else:
                print(f"    {shop_name}: 获取失败 - {result.get('msg')}")
                generator.add_data_row(shop_id, shop_name, date_range, {})
        except Exception as e:
            print(f"    {shop_name}: 请求异常 - {e}")
            generator.add_data_row(shop_id, shop_name, date_range, {})

    # 保存文件
    if not output_dir:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(output_dir, exist_ok=True)

    if is_multi_shop:
        output_file = f"{output_dir}/评论统计_多门店_{begin_date}_{end_date}.xlsx"
    else:
        output_file = f"{output_dir}/评论统计_{shop_ids[0]}_{begin_date}_{end_date}.xlsx"

    generator.save(output_file)
    return output_file


if __name__ == "__main__":
    """
    入口：直接运行 python excel_generator.py 即可生成报表

    使用方法：
        1. 修改 shop_config.json 中的 search_key（门店ID或名称）和日期
        2. 运行 python excel_generator.py

    search_key: 可以是单个门店ID、多个门店ID列表（JSON数组格式）或逗号分隔字符串
    platform: 0=全平台（默认），1=点评，2=美团
    """
    import json
    sys.path.insert(0, os.path.join(_CURRENT_DIR, '..', 'tuiguangtong'))
    from shop_search import resolve_shop

    print("=" * 50)
    print("评论统计报表生成")
    print("=" * 50)

    config_file = os.path.join(_CURRENT_DIR, 'shop_config.json')
    with open(config_file, encoding='utf-8') as f:
        config = json.load(f)

    search_key = config.get("search_key", "")
    platform = config.get("platform", 0)
    date_range = config.get("日期范围", {})

    if not search_key:
        print("错误: shop_config.json 中未设置 search_key")
        exit(1)

    # 解析 search_key（支持多种格式）
    if isinstance(search_key, list):
        shop_ids = search_key
    elif isinstance(search_key, str):
        if "," in search_key:
            shop_ids = [s.strip() for s in search_key.split(",")]
        else:
            shop_ids = [search_key]
    else:
        shop_ids = [str(search_key)]

    # 获取门店信息
    if len(shop_ids) == 1:
        shop_name, shop_id = resolve_shop(search_key if isinstance(search_key, str) else shop_ids[0], platform)
        print(f"已选择门店: {shop_name}")
    else:
        first_key = shop_ids[0]
        shop_name, _ = resolve_shop(first_key, platform)
        print(f"已选择门店数量: {len(shop_ids)}")

    platform_names = {0: "全平台", 1: "点评", 2: "美团"}
    print(f"平台: {platform_names.get(platform, '全平台')}")

    from review_statistics.client import ReviewStatisticsClient

    client = ReviewStatisticsClient()
    generate_report(
        client=client,
        shop_ids=shop_ids,
        begin_date=date_range.get("begin", ""),
        end_date=date_range.get("end", ""),
        platform=platform,
    )