"""
门店评论Excel生成器

=== 板块归属 ===
门店评论板块，对应目录 shop_review/

=== 样式说明 ===
- 第1行: 大标题，橙色背景
- 第2行: 表头，灰色背景
- 第3行起: 数据行，白色背景
- 字体: 微软雅黑
- 所有内容居中对齐
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, List

import sys
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)


# ============== 样式定义 ==============

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
DATA_FONT = Font(name="微软雅黑", size=10)

TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 平台名称映射
PLATFORM_NAMES = {1: "点评", 2: "美团"}


class ExcelGenerator:
    """门店评论Excel生成器"""

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "门店评论"

    def _get_headers(self):
        """获取表头列表"""
        return [
            "门店ID",
            "门店名称",
            "平台",
            "评论ID",
            "用户评分",
            "细致评分",
            "用户评论内容",
            "用户图片",
            "评论发布时间",
            "商户回复内容",
            "商户回复时间",
        ]

    def _get_widths(self):
        """获取列宽列表"""
        return [15, 30, 10, 25, 12, 39.125, 50, 40, 22, 50, 22]

    def _create_title_row(self, title: str = "门店评论"):
        """创建标题行（第1行）"""
        headers = self._get_headers()
        last_col = get_column_letter(len(headers))
        self.ws.merge_cells(f"A1:{last_col}1")
        title_cell = self.ws["A1"]
        title_cell.value = title
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = CENTER_ALIGN
        self.ws.row_dimensions[1].height = 28

    def _create_header_row(self):
        """创建表头行（第2行）"""
        headers = self._get_headers()
        widths = self._get_widths()

        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = self.ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width

        self.ws.row_dimensions[2].height = 22

    def add_review_row(self, review: Dict, platform_name: str = ""):
        """
        添加一条评论数据

        Args:
            review: 格式化后的评论数据
            platform_name: 平台名称（点评/美团）
        """
        row_data = [
            review.get("shop_id", ""),
            review.get("shop_name", ""),
            platform_name,
            review.get("review_id", ""),
            review.get("star", ""),
            review.get("score_detail", ""),
            review.get("content", ""),
            review.get("pictures", ""),
            review.get("create_time", ""),
            review.get("reply_content", ""),
            review.get("reply_time", ""),
        ]

        row_idx = self.ws.max_row + 1

        for col_idx, value in enumerate(row_data, start=1):
            cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 设置行高
        self.ws.row_dimensions[row_idx].height = 78

    def save(self, filepath: str):
        """保存Excel文件"""
        # 检查文件是否被占用
        if os.path.exists(filepath):
            try:
                with open(filepath, 'a'):
                    pass
            except IOError:
                print(f"错误: 文件被占用 {filepath}")
                print("请关闭 Excel 文件后重试")
                raise IOError(f"文件被占用: {filepath}")
        self.wb.save(filepath)
        print(f"Excel 文件已保存: {filepath}")


def generate_report(
    client,
    shop_ids: List[str],
    platform: int = 1,
    begin_date: str = None,
    end_date: str = None,
    output_dir: str = None
) -> str:
    """
    生成门店评论报表（支持多门店）

    Args:
        client: ShopReviewClient 实例
        shop_ids: 门店ID列表
        platform: 平台选择，1=点评，2=美团
        begin_date: 开始日期，格式 YYYY-MM-DD
        end_date: 结束日期，格式 YYYY-MM-DD
        output_dir: 输出目录，默认 D:\code\meituan\shop_review\reports

    Returns:
        生成的 Excel 文件路径
    """
    # 加载门店映射表
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    def get_shop_name(shop_id: str) -> str:
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}")
        return f"门店{shop_id}"

    # 设置默认日期
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    if not begin_date:
        begin_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")

    # 判断是否多门店
    is_multi_shop = len(shop_ids) > 1
    is_all_shops = len(shop_ids) > 50  # 超过50个门店视为全部门店模式

    # 创建Excel
    generator = ExcelGenerator()
    if is_all_shops:
        generator._create_title_row("门店评论（全部门店）")
    elif is_multi_shop:
        generator._create_title_row("门店评论（多门店）")
    else:
        generator._create_title_row("门店评论")
    generator._create_header_row()

    # 平台名称
    platform_name = PLATFORM_NAMES.get(platform, str(platform))

    # 全部门店模式：使用分页获取所有评论
    if is_all_shops:
        print(f"全部门店模式：{len(shop_ids)} 个门店，使用分页获取...")
        all_reviews = client.get_reviews_all_pages(shop_ids, str(platform), begin_date, end_date)

        # 添加评论数据
        for review in all_reviews:
            detail = client.get_review_detail(review)
            generator.add_review_row(detail, platform_name)

        total_count = len(all_reviews)

    else:
        # 遍历每个门店获取评论
        total_count = 0
        for shop_id in shop_ids:
            shop_name = get_shop_name(shop_id)
            print(f"  查询门店: {shop_name} ({platform_name})...")

            try:
                reviews = client.get_reviews(shop_id, str(platform), begin_date, end_date)
                print(f"    获取到 {len(reviews)} 条评论")
                total_count += len(reviews)

                # 添加评论数据
                for review in reviews:
                    detail = client.get_review_detail(review)
                    generator.add_review_row(detail, platform_name)

            except Exception as e:
                print(f"    获取失败: {e}")

    print(f"\n共获取到 {total_count} 条评论")

    # 保存文件
    if not output_dir:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(output_dir, exist_ok=True)

    if is_all_shops:
        filename = f"门店评论_全部门店_{platform_name}_{begin_date}_{end_date}.xlsx"
    elif is_multi_shop:
        filename = f"门店评论_多门店_{begin_date}_{end_date}.xlsx"
    else:
        filename = f"门店评论_{shop_ids[0]}_{platform_name}_{begin_date}_{end_date}.xlsx"

    filepath = os.path.join(output_dir, filename)
    generator.save(filepath)

    return filepath


if __name__ == "__main__":
    """
    入口：直接运行 python excel_generator.py 即可生成报表

    使用方法：
        1. 修改 shop_config.json 中的 search_key 和 platform
        2. 运行 python excel_generator.py

    shop_config.json 支持：
        - search_key: 单个门店ID、多个门店ID列表（JSON数组格式）
        - platform: 1=点评，2=美团（注意：本模块不支持 platform=0）
    """
    import json
    sys.path.insert(0, os.path.join(_CURRENT_DIR, '..', 'tuiguangtong'))
    from shop_search import resolve_shop
    from shop_review.client import ShopReviewClient, load_all_shop_ids

    print("=" * 50)
    print("门店评论报表生成")
    print("=" * 50)

    config_file = os.path.join(_CURRENT_DIR, 'shop_config.json')
    with open(config_file, encoding='utf-8') as f:
        config = json.load(f)

    search_key = config.get("search_key", "")
    platform = config.get("platform", 1)
    date_range = config.get("日期范围", {})
    begin_date = date_range.get("begin")
    end_date = date_range.get("end")

    # 解析 search_key（支持多种格式）
    if isinstance(search_key, list):
        shop_ids = search_key
        print(f"已选择门店数量: {len(shop_ids)}")
    elif isinstance(search_key, str):
        if search_key == "全部门店":
            # 从 all_shop_ids.json 加载所有门店
            shop_ids = load_all_shop_ids()
            print(f"已选择全部门店: {len(shop_ids)} 个")
        elif "," in search_key:
            shop_ids = [s.strip() for s in search_key.split(",")]
            print(f"已选择门店数量: {len(shop_ids)}")
        else:
            shop_ids = [search_key]
            print(f"已选择门店: {search_key}")
    else:
        shop_ids = [str(search_key)]
        print(f"已选择门店: {shop_ids[0]}")

    platform_names = {1: "点评", 2: "美团"}
    print(f"平台: {platform_names.get(platform, '未知')}")
    print(f"日期范围: {begin_date} 至 {end_date}")

    client = ShopReviewClient()

    generate_report(
        client=client,
        shop_ids=shop_ids,
        platform=platform,
        begin_date=begin_date,
        end_date=end_date,
    )