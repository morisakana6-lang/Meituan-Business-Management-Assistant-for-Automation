"""
交易分析多门店ID测试脚本
使用实际payload参数测试多个门店ID，生成汇总报表
"""

import sys
import os

_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)

from trade_analysis.client import TradeAnalysisClient


def test_multi_shop_report():
    """测试多个门店ID，生成汇总报表"""
    print("=" * 60)
    print("交易分析多门店ID测试")
    print("=" * 60)

    # 使用payload中的门店IDs
    shop_ids = ["1933643130", "57904793"]
    # 使用配置中的日期
    begin_date = "2026-05-08"
    end_date = "2026-05-14"
    platform = "0"

    print(f"\n测试门店IDs: {shop_ids}")
    print(f"日期范围: {begin_date} 至 {end_date}")
    print(f"平台: {platform}")

    client = TradeAnalysisClient()

    # 从shop_mapping获取门店名称
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    def get_shop_name(shop_id: str) -> str:
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}")
        return f"门店{shop_id}"

    results = []

    for shop_id in shop_ids:
        print(f"\n正在查询门店 {shop_id}...")

        try:
            metrics = client.get_metrics(
                shop_id=shop_id,
                platform=platform,
                begin_date=begin_date,
                end_date=end_date
            )

            shop_name = get_shop_name(shop_id)
            results.append({
                "shop_id": shop_id,
                "shop_name": shop_name,
                "metrics": metrics,
            })
            print(f"  {shop_name}: 获取成功")
            for key, value in metrics.items():
                if not key.endswith("_suffix"):
                    print(f"    {key}: {value}")

        except Exception as e:
            print(f"  请求异常: {e}")

    # 生成Excel报表
    if results:
        generate_excel_report(results, begin_date, end_date)


def generate_excel_report(results, begin_date, end_date):
    """生成Excel报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
    DATA_FONT = Font(name="微软雅黑", size=10)

    TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "交易分析报表"

    # 表头
    headers = [
        "门店ID", "推广门店", "时间",
        "下单人数（人）", "下单券数（张）", "下单金额（原价）（元）",
        "下单金额（元）", "核销人数（人）", "核销券数（张）",
        "核销金额（原价）（元）", "核销金额（元）",
        "退款券数（张）", "退款金额（原价）（元）"
    ]

    widths = [17.25, 28.875, 22.0, 14.0, 14.0, 18.0, 15.0,
              14.0, 14.0, 18.0, 15.0, 14.0, 18.0]

    # 标题行
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value = "交易分析报表（多门店）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28

    # 表头行
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # 数据行
    date_range = f"{begin_date}至{end_date}"

    for row_idx, result in enumerate(results, start=3):
        shop_id = result["shop_id"]
        shop_name = result["shop_name"]
        metrics = result["metrics"]

        def get_val(key):
            val = metrics.get(key, "0")
            if isinstance(val, str):
                # 移除逗号和货币符号
                val = val.replace(",", "").replace("¥", "").replace("$", "")
            try:
                return float(val)
            except:
                return 0.0

        row_data = [
            shop_id,
            shop_name,
            date_range,
            get_val("order_person_count"),
            get_val("order_ticket_count"),
            get_val("order_original_amount"),
            get_val("order_amount"),
            get_val("redeem_person_count"),
            get_val("redeem_ticket_count"),
            get_val("redeem_original_amount"),
            get_val("redeem_amount"),
            get_val("refund_ticket_count"),
            get_val("refund_original_amount"),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        ws.row_dimensions[row_idx].height = 18

    # 保存
    output_dir = os.path.join(_CURRENT_DIR, "reports")
    os.makedirs(output_dir, exist_ok=True)
    output_file = f"{output_dir}/交易分析_多门店_{begin_date}_{end_date}.xlsx"

    # 检查文件是否被占用
    if os.path.exists(output_file):
        try:
            with open(output_file, 'a'):
                pass
        except IOError:
            print(f"错误: 文件被占用 {output_file}")
            print("请关闭 Excel 文件后重试")
            raise IOError(f"文件被占用: {output_file}")

    wb.save(output_file)
    print(f"\nExcel 文件已保存: {output_file}")


if __name__ == "__main__":
    test_multi_shop_report()
