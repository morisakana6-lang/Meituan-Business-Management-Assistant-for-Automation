"""
客流分析Excel生成器

=== 主要作用 ===
将美团经营宝客流分析的指标数据生成 Excel 汇总报表文件。

=== 板块归属 ===
客流分析板块，对应目录 customer_flow/

=== 输出格式 ===
39列，7个分组：
| 分组 | 列范围 | 字段数 | 字段 |
|------|--------|--------|------|
| 基本信息 | A-D | 4 | 门店ID、推广门店、门店所在城市、时间 |
| 客流统计报表 | E-P | 12 | 曝光人数、曝光次数、访问人数、访问次数、曝光访问转化率、意向转化人数、意向转化率、下单人数、留资人数、累计收藏人数、新增收藏人数、新增打卡人数 |
| 引流用户数据 | Q-S | 3 | 美团引流顾客，自然到店顾客、潜在顾客 |
| 在线咨询 | T-V | 3 | 在线咨询人数、在线咨询留资数、咨询留资转化率 |
| 门店交易情况 | W-AF | 10 | 下单人数、下单券数、下单金额（原价）、核销人数、核销券数、核销金额（原价）、退款券数、退款金额（原价） |
| 门店评价概览 | AG-AK | 5 | 新增评价数、新增差评数、差评回复率、新增好评数、累计评价数 |
| 星级概览 | AL-AM | 2 | 点评星级、美团星级 |

=== 样式说明 ===
- 第1行: 大标题，橙色背景
- 第2行: 表头，灰色背景
- 第3行起: 数据行，白色背景
- 字体: 微软雅黑
- 全部居中对齐
"""

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, List

import sys
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)


# ============== 样式定义 ==============

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
DATA_FONT = Font(name="微软雅黑", size=10)

TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


class ExcelGenerator:
    """客流分析Excel生成器"""

    # 分组定义：(列范围, 标题名称, 背景颜色)
    GROUPS = [
        (1, 16, "客流统计报表", "FFFBE5D5"),      # 浅橙
        (17, 19, "引流用户数据情况", "FFD6E8FF"),   # 浅蓝
        (20, 22, "在线咨询", "FFD5F5DC"),          # 浅绿
        (23, 32, "门店交易情况", "FFF0E68C"),       # 浅黄
        (33, 37, "门店评价概览", "FFEEE8F5"),       # 浅紫
        (38, 39, "星级概览", "FFFCE4EC"),          # 浅粉
    ]

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "付费版线上数据"

    def _create_title_row(self):
        """创建分组标题行（第1行）"""
        # 创建分组标题
        for start_col, end_col, group_name, color in self.GROUPS:
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)

            # 合并单元格
            if start_col == end_col:
                cell_ref = f"{start_letter}1"
            else:
                cell_ref = f"{start_letter}1:{end_letter}1"

            self.ws.merge_cells(cell_ref)

            # 设置分组标题
            cell = self.ws.cell(row=1, column=start_col)
            cell.value = group_name
            cell.font = TITLE_FONT
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = CENTER_ALIGN

        self.ws.row_dimensions[1].height = 28

    def _create_header_row(self):
        """创建表头行（第2行）"""
        headers = self._get_headers()
        widths = self._get_widths()

        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = self.ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width

        self.ws.row_dimensions[2].height = 22

    def _get_headers(self):
        """获取表头列表（40列）"""
        return [
            # 基本信息 (A-D, 4列) - 合并到客流统计报表分组
            "门店ID",
            "推广门店",
            "门店所在城市",
            "时间",
            # 客流统计报表 (E-P, 12列)
            "曝光人数（人）",
            "曝光次数（次）",
            "访问人数（人）",
            "访问次数（次）",
            "曝光访问转化率（%）",
            "意向转化人数（人）",
            "意向转化率（%）",
            "下单人数（人）",
            "留资人数（人）",
            "累计收藏人数（人）",
            "新增收藏人数（人）",
            "新增打卡人数（人）",
            # 引流用户数据 (Q-S, 3列)
            "美团引流顾客（人）",
            "自然到店顾客（人）",
            "潜在顾客（人）",
            # 在线咨询 (T-V, 3列)
            "在线咨询人数（人）",
            "在线咨询留资数（人）",
            "咨询留资转化率（%）",
            # 门店交易情况 (W-AF, 10列)
            "下单人数（人）",
            "下单券数（张）",
            "下单金额（原价）（元）",
            "下单金额（元）",
            "核销人数（人）",
            "核销券数（张）",
            "核销金额（原价）（元）",
            "核销金额（元）",
            "退款券数（张）",
            "退款金额（原价）（元）",
            # 门店评价概览 (AG-AK, 5列)
            "新增评价数",
            "新增差评数",
            "差评回复率",
            "新增好评数",
            "累计评价数",
            # 星级概览 (AL-AM, 2列)
            "点评星级",
            "美团星级",
        ]

    def _get_widths(self):
        """获取列宽列表（40列）"""
        return [
            # 基本信息 (4列)
            22.75, 30, 15, 22,
            # 客流统计报表 (12列)
            15, 15, 15, 15, 18, 16, 15, 13, 13, 16, 16, 15,
            # 引流用户数据 (3列)
            18, 18, 15,
            # 在线咨询 (3列)
            18, 18, 18,
            # 门店交易情况 (10列)
            13, 13, 20, 15, 13, 13, 20, 15, 13, 20,
            # 门店评价概览 (5列)
            12, 12, 12, 12, 12,
            # 星级概览 (2列)
            12, 12,
        ]

    def _get_metric_value_float(self, metrics: Dict, metric_id: str) -> float:
        """从指标字典中获取指定指标的值（转换为浮点数）"""
        value = metrics.get(metric_id, "0")
        if isinstance(value, str):
            value = value.replace(",", "").replace("￥", "").replace("¥", "").replace("$", "").replace("%", "").replace(" ", "")
            if "万" in value:
                value = value.replace("万", "")
                try:
                    return float(value) * 10000
                except (ValueError, TypeError):
                    return 0.0
            elif "亿" in value:
                value = value.replace("亿", "")
                try:
                    return float(value) * 100000000
                except (ValueError, TypeError):
                    return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def add_data_row(
        self,
        shop_id: str,
        shop_name: str,
        shop_city: str,
        date_range: str,
        flow_metrics: Dict,
        user_data_metrics: Dict,
        consultation_metrics: Dict,
        trade_metrics: Dict,
        review_metrics: Dict,
        star_metrics: Dict,
    ):
        """
        添加一行数据

        Args:
            shop_id: 门店ID
            shop_name: 门店名称
            shop_city: 门店所在城市
            date_range: 时间范围
            flow_metrics: 客流统计数据
            user_data_metrics: 引流用户数据
            consultation_metrics: 在线咨询数据
            trade_metrics: 交易分析数据
            review_metrics: 评论统计数据
            star_metrics: 星级评分数据
        """
        row_data = [
            # 基本信息 (4列)
            shop_id,
            shop_name,
            shop_city,
            date_range,
            # 客流统计报表 (12列)
            self._get_metric_value_float(flow_metrics, "exposure_count"),
            self._get_metric_value_float(flow_metrics, "exposure_times"),
            self._get_metric_value_float(flow_metrics, "visitor_count"),
            self._get_metric_value_float(flow_metrics, "visit_times"),
            self._get_metric_value_float(flow_metrics, "exposure_visit_rate"),
            self._get_metric_value_float(flow_metrics, "intention_count"),
            self._get_metric_value_float(flow_metrics, "intention_rate"),
            self._get_metric_value_float(flow_metrics, "order_count"),
            self._get_metric_value_float(flow_metrics, "retention_count"),
            self._get_metric_value_float(flow_metrics, "favorites_count"),
            self._get_metric_value_float(flow_metrics, "new_favorites_count"),
            self._get_metric_value_float(flow_metrics, "new_checkin_count"),
            # 引流用户数据 (3列)
            self._get_metric_value_float(user_data_metrics, "traffic_user_count"),
            self._get_metric_value_float(user_data_metrics, "natural_user_count"),
            self._get_metric_value_float(user_data_metrics, "potential_user_count"),
            # 在线咨询 (3列)
            self._get_metric_value_float(consultation_metrics, "在线咨询人数"),
            self._get_metric_value_float(consultation_metrics, "在线咨询留资数"),
            consultation_metrics.get("咨询留资转化率", ""),
            # 门店交易情况 (10列)
            self._get_metric_value_float(trade_metrics, "order_person_count"),
            self._get_metric_value_float(trade_metrics, "order_ticket_count"),
            self._get_metric_value_float(trade_metrics, "order_original_amount"),
            self._get_metric_value_float(trade_metrics, "order_amount"),
            self._get_metric_value_float(trade_metrics, "redeem_person_count"),
            self._get_metric_value_float(trade_metrics, "redeem_ticket_count"),
            self._get_metric_value_float(trade_metrics, "redeem_original_amount"),
            self._get_metric_value_float(trade_metrics, "redeem_amount"),
            self._get_metric_value_float(trade_metrics, "refund_ticket_count"),
            self._get_metric_value_float(trade_metrics, "refund_original_amount"),
            # 门店评价概览 (5列)
            self._get_metric_value_float(review_metrics, "新增评价数"),
            self._get_metric_value_float(review_metrics, "新增差评数"),
            review_metrics.get("差评回复率", ""),
            self._get_metric_value_float(review_metrics, "新增好评数"),
            self._get_metric_value_float(review_metrics, "累计评价数"),
            # 星级概览 (2列)
            star_metrics.get("点评星级", ""),
            star_metrics.get("美团星级", ""),
        ]

        row_idx = self.ws.max_row + 1

        for col_idx, value in enumerate(row_data, start=1):
            cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        self.ws.row_dimensions[row_idx].height = 18

    def save(self, filepath: str):
        """保存Excel文件"""
        if os.path.exists(filepath):
            try:
                with open(filepath, 'a'):
                    pass
            except IOError:
                print(f"错误: 文件被占用 {filepath}")
                print("请关闭 Excel 文件后重试")
                raise IOError(f"文件被占用: {filepath}")
        self.wb.save(filepath)
        print(f"Excel 文件已保存: {filepath}")


def generate_report(client, shop_ids: list, begin_date: str, end_date: str, platform: int = 0):
    """
    生成汇总报表（支持单门店和多门店）

    Args:
        client: CustomerFlowClient 实例
        shop_ids: 门店ID列表，支持单个或多个门店
        begin_date: 开始日期，格式 YYYY-MM-DD
        end_date: 结束日期，格式 YYYY-MM-DD
        platform: 平台，0=全平台，1=点评，2=美团

    Returns:
        生成的 Excel 文件路径
    """
    from customer_flow.client import get_shop_info
    from tuiguangtong.shop_search import load_mapping

    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    def get_shop_name(shop_id: str) -> str:
        if shop_id == "0":
            return "全部门店汇总数据"
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}")
        return f"门店{shop_id}"

    generator = ExcelGenerator()

    is_multi_shop = len(shop_ids) > 1
    generator._create_title_row()  # 创建分组标题行
    generator._create_header_row()  # 创建表头行

    date_range = f"{begin_date}至{end_date}"
    print(f"开始生成客流分析报表: {begin_date} ~ {end_date}")

    for shop_id in shop_ids:
        shop_name = get_shop_name(shop_id)
        shop_info = get_shop_info(shop_id)
        shop_city = shop_info.get("city", "")
        print(f"  查询门店: {shop_name} ({shop_city})...")

        try:
            # 获取客流统计数据
            flow_metrics = client.get_metrics(
                shop_id=shop_id,
                platform=str(platform),
                begin_date=begin_date,
                end_date=end_date
            )

            # 获取评论统计数据（旧版API）
            review_metrics = client.get_review_statistics(
                shop_id=shop_id,
                begin_date=begin_date,
                end_date=end_date,
                platform=str(platform)
            )

            # 获取星级评分数据
            star_metrics = client.get_star_rating(
                shop_id=shop_id,
                begin_date=begin_date,
                end_date=end_date,
                platform=str(platform)
            )

            # 获取在线咨询数据
            consultation_metrics = client.get_online_consultation(
                shop_id=shop_id,
                begin_date=begin_date,
                end_date=end_date,
                platform=str(platform)
            )

            # 获取交易分析数据
            trade_metrics = client.get_trade_analysis(
                shop_id=shop_id,
                begin_date=begin_date,
                end_date=end_date,
                platform=str(platform)
            )

            # 从 flow_metrics 中提取引流用户数据
            user_data_metrics = {
                "traffic_user_count": flow_metrics.get("traffic_user_count", "0"),
                "natural_user_count": flow_metrics.get("natural_user_count", "0"),
                "potential_user_count": flow_metrics.get("potential_user_count", "0"),
            }

            generator.add_data_row(
                shop_id=shop_id,
                shop_name=shop_name,
                shop_city=shop_city,
                date_range=date_range,
                flow_metrics=flow_metrics,
                user_data_metrics=user_data_metrics,
                consultation_metrics=consultation_metrics,
                trade_metrics=trade_metrics,
                review_metrics=review_metrics,
                star_metrics=star_metrics,
            )
            print(f"    {shop_name}: 获取成功")

        except Exception as e:
            error_msg = str(e)
            if "noRightsShop" in error_msg:
                print(f"    {shop_name}: 无权限访问，已跳过")
            else:
                print(f"    {shop_name}: 获取失败 - {e}")
                generator.add_data_row(
                    shop_id=shop_id,
                    shop_name=shop_name,
                    shop_city=shop_city,
                    date_range=date_range,
                    flow_metrics={},
                    user_data_metrics={},
                    consultation_metrics={},
                    trade_metrics={},
                    review_metrics={},
                    star_metrics={},
                )

    output_dir = os.path.join(_CURRENT_DIR, "reports")
    os.makedirs(output_dir, exist_ok=True)

    if is_multi_shop:
        output_file = f"{output_dir}/客流分析_多门店_{begin_date}_{end_date}.xlsx"
    else:
        output_file = f"{output_dir}/客流分析_{shop_ids[0]}_{begin_date}_{end_date}.xlsx"

    generator.save(output_file)
    return output_file


if __name__ == "__main__":
    """
    入口：直接运行 python excel_generator.py 即可生成报表

    使用方法：
        1. 修改 shop_config.json 中的 search_key（门店ID或名称）和日期
        2. 运行 python excel_generator.py

    search_key: 可以是单个门店ID、多个门店ID列表（JSON数组格式）或逗号分隔字符串
    platform: 0=全平台（默认），1=点评，2=美团
    """
    import json
    sys.path.insert(0, os.path.join(_CURRENT_DIR, '..', 'tuiguangtong'))
    from shop_search import resolve_shop

    config_path = os.path.join(_CURRENT_DIR, "shop_config.json")
    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    search_key = config.get("search_key", "")
    platform = config.get("platform", 0)
    date_range = config["日期范围"]

    if not search_key:
        print("错误: shop_config.json 中未设置 search_key")
        exit(1)

    if isinstance(search_key, list):
        shop_ids = search_key
    elif isinstance(search_key, str):
        if "," in search_key:
            shop_ids = [s.strip() for s in search_key.split(",")]
        else:
            shop_ids = [search_key]
    else:
        shop_ids = [str(search_key)]

    if len(shop_ids) == 1:
        shop_name, shop_id = resolve_shop(search_key if isinstance(search_key, str) else shop_ids[0], platform)
        print(f"已选择门店: {shop_name}")
    else:
        first_key = shop_ids[0]
        shop_name, _ = resolve_shop(first_key, platform)
        print(f"已选择门店数量: {len(shop_ids)}")

    platform_names = {0: "全平台", 1: "点评", 2: "美团"}
    print(f"平台: {platform_names.get(platform, '全平台')}")

    from customer_flow.client import CustomerFlowClient

    client = CustomerFlowClient()
    generate_report(
        client=client,
        shop_ids=shop_ids,
        begin_date=date_range["begin"],
        end_date=date_range["end"],
        platform=platform,
    )