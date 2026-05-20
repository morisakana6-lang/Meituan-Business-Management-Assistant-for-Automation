"""
Excel公共样式
"""

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# 表头样式
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# 单元格样式
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 汇总行样式
SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SUMMARY_FONT = Font(bold=True, size=11)
