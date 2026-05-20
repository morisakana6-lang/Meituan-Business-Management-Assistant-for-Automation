"""
全站推广多门店ID测试脚本
测试多个门店ID，生成汇总报表（每行一个门店）
"""

import sys
import os

_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)

from quantianzhan.client import QuantianzhanClient


def test_multi_shop_report():
    """测试多个门店ID，生成汇总报表"""
    print("=" * 60)
    print("全站推广多门店ID测试")
    print("=" * 60)

    # 测试用的多个门店ID
    shop_ids = ["18666748", "835187138"]

    print(f"\n测试门店IDs: {shop_ids}")
    print(f"日期范围: 2026-02-01 至 2026-03-14")

    client = QuantianzhanClient()
    platform = 0

    # 获取shop_mapping来获取门店名称
    from tuiguangtong.shop_search import load_mapping
    mapping_data = load_mapping()
    shops = mapping_data.get("shops", [])

    # 构建 id -> name 映射
    def get_shop_name(shop_id: str) -> str:
        for shop in shops:
            for id_info in shop.get("ids", []):
                if id_info.get("id") == shop_id:
                    return shop.get("name", f"门店{shop_id}")
        return f"门店{shop_id}"

    results = []

    for shop_id in shop_ids:
        print(f"\n正在查询门店 {shop_id}...")

        try:
            result = client.get_report(
                begin_date="2026-02-01",
                end_date="2026-03-14",
                shop_ids=shop_id,
                platform=platform,
            )

            if result.get('code') == 200:
                msg = result.get('msg', {})
                total = msg.get('total', [])

                # 提取指标
                metrics = {}
                for item in total:
                    metric_id = item.get('id', '')
                    value = item.get('value', '0')
                    metrics[metric_id] = value

                shop_name = get_shop_name(shop_id)
                results.append({
                    "shop_id": shop_id,
                    "shop_name": shop_name,
                    "metrics": metrics,
                })
                print(f"  {shop_name}: 获取成功")
            else:
                print(f"  API错误: {result.get('msg')}")

        except Exception as e:
            print(f"  请求异常: {e}")

    # 生成Excel报表
    if results:
        generate_excel_report(results, platform)


def generate_excel_report(results, platform):
    """生成Excel报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
    DATA_FONT = Font(name="微软雅黑", size=10)

    TITLE_FILL = PatternFill(start_color="FFFBE5D5", end_color="FFFBE5D5", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "全站推广报表"

    # 表头
    headers = ["门店ID", "门店名称", "花费(元)", "现金花费(元)", "全站浏览量(次)", "下单券数(张)", "每张券成本(元)"]

    widths = [15, 28, 12, 14, 16, 14, 14]

    # 标题行
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value = "全站推广报表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28

    # 表头行
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # 数据行
    for row_idx, result in enumerate(results, start=3):
        shop_id = result["shop_id"]
        shop_name = result["shop_name"]
        metrics = result["metrics"]

        def get_val(metric_id):
            val = metrics.get(metric_id, "0").replace(",", "")
            try:
                return float(val)
            except:
                return 0.0

        cost = get_val("T30001")
        cash_cost = get_val("T30047")
        pv = get_val("T500001")
        coupon_count = get_val("T600008")
        coupon_cost = get_val("T600009")

        row_data = [
            shop_id,
            shop_name,
            f"{cost:.2f}",
            f"{cash_cost:.2f}",
            str(int(pv)),
            str(int(coupon_count)),
            f"{coupon_cost:.2f}",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = WHITE_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        ws.row_dimensions[row_idx].height = 18

    # 保存
    output_dir = os.path.join(_CURRENT_DIR, "reports")
    os.makedirs(output_dir, exist_ok=True)
    output_file = f"{output_dir}/全站推广_多门店_2026-02-01_2026-03-14.xlsx"
    wb.save(output_file)
    print(f"\nExcel 文件已保存: {output_file}")


if __name__ == "__main__":
    test_multi_shop_report()
