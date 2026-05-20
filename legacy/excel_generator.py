"""
美团经营宝数据 Excel 生成模块
将报表数据生成符合目标样式的 Excel 文件
"""

import json
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Excel 生成器"""

    # 列定义
    COLUMNS = [
        {"header": "日期", "key": "date", "width": 12},
        {"header": "门店名称", "key": "shop_name", "width": 25},
        {"header": "门店ID", "key": "shop_id", "width": 15},
        {"header": "花费(元)", "key": "T30001", "width": 12},
        {"header": "现金花费(元)", "key": "T30047", "width": 12},
        {"header": "曝光(次)", "key": "T30002", "width": 12},
        {"header": "点击(次)", "key": "T30003", "width": 10},
        {"header": "点击均价(元)", "key": "cpc", "width": 12},
        {"header": "商户浏览量(次)", "key": "T30005", "width": 14},
        {"header": "查看图片(次)", "key": "T30006", "width": 12},
        {"header": "查看评论(次)", "key": "T30007", "width": 12},
        {"header": "查看店铺信息(次)", "key": "T30039", "width": 14},
        {"header": "查看团购(次)", "key": "T30009", "width": 12},
        {"header": "感兴趣(次)", "key": "T30083", "width": 12},
        {"header": "团购订单量(个)", "key": "T30020", "width": 14},
        {"header": "订单量(个)", "key": "T30049", "width": 12},
    ]

    # 样式定义
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 需要汇总的数值列（排除日期、门店名称、门店ID）
    SUM_COLUMNS = ["T30001", "T30047", "T30002", "T30003", "cpc", "T30005", "T30006", "T30007", "T30039", "T30009", "T30083", "T30020", "T30049"]

    def __init__(self, shop_name: str = "丽减美瘦吧·减肥瘦身(新澳城店)", shop_id: str = "764510450"):
        self.shop_name = shop_name
        self.shop_id = shop_id
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "经营报表"
        self.data_rows = []  # 存储所有数据行用于汇总

    def _create_header(self):
        """创建表头"""
        for col_idx, col_def in enumerate(self.COLUMNS, start=1):
            cell = self.ws.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 设置列宽
        for col_idx, col_def in enumerate(self.COLUMNS, start=1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    def _get_metric_value(self, metrics: list, metric_id: str) -> str:
        """从指标列表中获取指定指标的值（保持原格式，含逗号）"""
        for m in metrics:
            if m.get("id") == metric_id:
                return m.get("value", "0")
        return "0"

    def _get_metric_value_float(self, metrics: list, metric_id: str) -> float:
        """从指标列表中获取指定指标的值（转换为浮点数）"""
        for m in metrics:
            if m.get("id") == metric_id:
                value = m.get("value", "0")
                if isinstance(value, str):
                    # 先移除逗号、货币符号、百分号和空格
                    value = value.replace(",", "").replace("￥", "").replace("¥", "").replace("$", "").replace("%", "").replace(" ", "")
                    # 处理带"万"、"亿"的单位
                    if "万" in value:
                        value = value.replace("万", "")
                        try:
                            return float(value) * 10000
                        except (ValueError, TypeError):
                            return 0.0
                    elif "亿" in value:
                        value = value.replace("亿", "")
                        try:
                            return float(value) * 100000000
                        except (ValueError, TypeError):
                            return 0.0
                try:
                    return float(value)
                except (ValueError, TypeError):
                    return 0.0
        return 0.0

    def _calculate_cpc(self, metrics: list) -> str:
        """计算点击均价 CPC = 花费 / 点击"""
        cost_val = self._get_metric_value_float(metrics, "T30001")
        clicks_val = self._get_metric_value_float(metrics, "T30003")

        try:
            if clicks_val > 0:
                return f"{cost_val / clicks_val:.2f}"
        except (ValueError, ZeroDivisionError):
            pass
        return "0.00"

    def add_data_row(self, date: str, metrics: list):
        """添加一行数据"""
        row_data = {
            "date": date,
            "shop_name": self.shop_name,
            "shop_id": self.shop_id,
            "T30001": self._get_metric_value(metrics, "T30001"),
            "T30047": self._get_metric_value(metrics, "T30047"),
            "T30002": self._get_metric_value(metrics, "T30002"),
            "T30003": self._get_metric_value(metrics, "T30003"),
            "cpc": self._calculate_cpc(metrics),
            "T30005": self._get_metric_value(metrics, "T30005"),
            "T30006": self._get_metric_value(metrics, "T30006"),
            "T30007": self._get_metric_value(metrics, "T30007"),
            "T30039": self._get_metric_value(metrics, "T30039"),
            "T30009": self._get_metric_value(metrics, "T30009"),
            "T30083": self._get_metric_value(metrics, "T30083"),
            "T30020": self._get_metric_value(metrics, "T30020"),
            "T30049": self._get_metric_value(metrics, "T30049"),
        }

        # 存储数据用于汇总
        self.data_rows.append(row_data)

        row_idx = self.ws.max_row + 1
        for col_idx, col_def in enumerate(self.COLUMNS, start=1):
            key = col_def["key"]
            value = row_data.get(key, "")

            cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = self.THIN_BORDER

            # 数字右对齐，日期和文本左对齐
            if key in ["date", "shop_name", "shop_id"]:
                cell.alignment = self.CELL_ALIGNMENT
            else:
                cell.alignment = self.NUMBER_ALIGNMENT

    def _add_summary_row(self):
        """添加汇总行"""
        # 计算各列总和
        sums = {}
        for col in self.SUM_COLUMNS:
            total = 0.0
            for row in self.data_rows:
                try:
                    # 处理带逗号的数字字符串，如 "146,504" -> 146504
                    value_str = row.get(col, "0").replace(",", "")
                    total += float(value_str)
                except (ValueError, TypeError):
                    pass
            sums[col] = total

        # 计算总点击均价 CPC = 总花费 / 总点击
        total_cost = sums.get("T30001", 0)
        total_clicks = sums.get("T30003", 0)
        if total_clicks > 0:
            sums["cpc"] = total_cost / total_clicks
        else:
            sums["cpc"] = 0

        # 汇总行数据
        summary_data = {
            "date": "汇总",
            "shop_name": self.shop_name,
            "shop_id": self.shop_id,
            **{col: sums.get(col, 0) for col in self.SUM_COLUMNS}
        }

        # 设置 CPC 格式
        summary_data["cpc"] = f"{sums['cpc']:.2f}" if isinstance(sums.get('cpc'), float) else sums.get('cpc', '0.00')
        # 其他数值列格式
        for col in self.SUM_COLUMNS:
            if col != "cpc" and isinstance(summary_data[col], float):
                if col in ["T30001", "T30047", "cpc"]:
                    summary_data[col] = f"{summary_data[col]:.2f}"
                else:
                    summary_data[col] = str(int(summary_data[col])) if summary_data[col] == int(summary_data[col]) else f"{summary_data[col]:.2f}"

        row_idx = self.ws.max_row + 1

        # 汇总行样式
        SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄色
        SUMMARY_FONT = Font(bold=True, size=11)

        for col_idx, col_def in enumerate(self.COLUMNS, start=1):
            key = col_def["key"]
            value = summary_data.get(key, "")

            cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = self.THIN_BORDER
            cell.fill = SUMMARY_FILL
            cell.font = SUMMARY_FONT

            if key in ["date", "shop_name", "shop_id"]:
                cell.alignment = self.CELL_ALIGNMENT
            else:
                cell.alignment = self.NUMBER_ALIGNMENT

        return summary_data

    def save(self, filepath: str):
        """保存 Excel 文件"""
        # 添加汇总行
        self._add_summary_row()
        self.wb.save(filepath)
        print(f"Excel 文件已保存: {filepath}")


def generate_daily_report(client, begin_date: str, end_date: str, shop_ids: str, shop_name: str, shop_id: str):
    """
    生成每日报表 Excel

    Args:
        client: MeituanReportClient 实例
        begin_date: 开始日期 YYYY-MM-DD
        end_date: 结束日期 YYYY-MM-DD
        shop_ids: 门店ID
        shop_name: 门店名称
        shop_id: 门店ID字符串
    """
    generator = ExcelGenerator(shop_name=shop_name, shop_id=shop_id)
    generator._create_header()

    # 解析日期
    start = datetime.strptime(begin_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")

    current = start
    total_days = (end - start).days + 1
    current_day = 0

    print(f"开始生成 {begin_date} ~ {end_date} 的每日报表，共 {total_days} 天...")

    while current <= end:
        date_str = current.strftime("%Y-%m-%d")
        date_display = current.strftime("%m-%d")  # 格式化为 MM-DD
        current_day += 1

        print(f"  [{current_day}/{total_days}] 查询 {date_str}...", end=" ")

        try:
            metrics = client.get_metrics(date_str, date_str, shop_ids)
            generator.add_data_row(date_display, metrics)
            print(f"OK (花费={generator._get_metric_value(metrics, 'T30001')})")
        except Exception as e:
            print(f"失败: {e}")
            # 尝试重新获取凭证后重试
            if "401" in str(e) or "403" in str(e) or "Unauthorized" in str(e):
                print("    检测到认证错误，可能是 Cookie/mtgsig 过期")
                print("    请运行: python get_credentials.py --update")
            # 添加空行保持日期连续
            generator.add_data_row(date_display, [])

        # 随机延迟，避免高频请求
        import random
        import time
        delay = random.uniform(1, 2)
        time.sleep(delay)

        current += timedelta(days=1)

    # 保存文件
    output_file = f"report_{shop_id}_{begin_date}_{end_date}.xlsx"
    generator.save(output_file)
    return output_file


if __name__ == "__main__":
    # 测试用
    print("此模块需配合 meituan_client.py 使用")
    print("使用方法: from excel_generator import generate_daily_report")